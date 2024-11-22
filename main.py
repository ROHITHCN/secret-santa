import openpyxl
import pandas as pd
import random

class SecretSanta:
    def __init__(self, employees, previous_assignments):
        self.employees = employees
        self.previous_assignments = previous_assignments

    def generate_derangement(self, n):
        """
        Generates a derangement for numbers from 0 to n-1.
        """
        derangement = list(range(n))
        while True:
            random.shuffle(derangement)
            if all(i != derangement[i] for i in range(n)):
                return derangement

    def assign_secret_santas(self):
        # Map employee email to their previous Secret Santa child for easy lookup
        prev_assignments_dict = {
            assignment['employee_email']: assignment['secret_child_email']
            for assignment in self.previous_assignments
        }

        employee_emails = [employee['employee_email'] for employee in self.employees]
        n = len(employee_emails)

        for _ in range(100):  # Retry loop to handle conflicts with previous year
            # Generate a derangement for indices
            derangement = self.generate_derangement(n)
            
            # Check if the derangement conflicts with last year's assignments
            conflict_found = False
            for i, email in enumerate(employee_emails):
                if prev_assignments_dict.get(email) == employee_emails[derangement[i]]:
                    conflict_found = True
                    break

            if not conflict_found:
                # Create the assignment based on the derangement
                assignments = []
                for i, email in enumerate(employee_emails):
                    current_employee = next(emp for emp in self.employees if emp['employee_email'] == email)
                    secret_child_email = employee_emails[derangement[i]]
                    secret_child = next(emp for emp in self.employees if emp['employee_email'] == secret_child_email)
                    
                    assignments.append({
                        'employee_name': current_employee['employee_name'],
                        'employee_email': current_employee['employee_email'],
                        'secret_child_name': secret_child['employee_name'],
                        'secret_child_email': secret_child['employee_email']
                    })
                return assignments  # Return the successful assignments

        # If no valid assignment is found after retries
        raise Exception("Failed to generate valid Secret Santa assignments after multiple attempts!")

def read_employees_from_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    employees = [
        {'employee_name': row[0], 'employee_email': row[1]}
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]
    return employees

def read_previous_assignments_from_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    previous_assignments = [
        {'employee_name': row[0], 'employee_email': row[1], 'secret_child_name': row[2], 'secret_child_email': row[3]}
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]
    return previous_assignments

def write_assignments_to_csv(assignments, output_file_path):
    df = pd.DataFrame(assignments)
    df.to_csv(output_file_path, index=False)
    print(f"Assignments written to {output_file_path}")

# Main execution
def main():
    employees = read_employees_from_xlsx('data/employee.xlsx')
    previous_assignments = read_previous_assignments_from_xlsx('data/previous_assignments.xlsx')
    
    secret_santa = SecretSanta(employees, previous_assignments)
    assignments = secret_santa.assign_secret_santas()

    write_assignments_to_csv(assignments, 'data/output.csv')

if __name__ == "__main__":
    main()
